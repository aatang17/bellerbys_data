"""
Load Bellerbys SAPM student list from Excel.
Returns students grouped by pathway with subject column names.
Can optionally read current grades from Excel (default: BNBU SAPM - Semester 1 Grades_v2.xlsx).
Ensures student_name is unique across the cohort (e.g. two "Iris" become "Iris Yang" and "Iris Chen")
so offer matching by name attaches to the correct person.
"""
import os
from collections import Counter
from typing import Optional

STUDENT_ID_COL  = 0
FIRST_NAME_COL  = 1
LAST_NAME_COL   = 2
ENGLISH_NAME_COL = 3
PATHWAY_COL     = 7

PATHWAY_SUBJECTS = {
    "Business Mgmt": ["AES", "Business Management", "Statistics with Project Skills", "Economics"],
    "Media":         ["AES", "Media Analysis", "Film Production", "Print & Digital News"],
    "Computing":     ["AES", "Mathematics", "Physics", "Statistics with Project Skills"],
}

PATHWAY_ALIASES = {
    "Biz Mgt": "Business Mgmt",
    "Business Mgmt": "Business Mgmt",
    "Business Management": "Business Mgmt",
    "Business": "Business Mgmt",
    "Media": "Media",
    "Computing": "Computing",
}


def _pathway_from_raw(raw: str) -> str | None:
    """Map pathway column value to one of Business Mgmt, Media, Computing. Case-insensitive; accepts partial matches."""
    if not raw or not raw.strip():
        return None
    r = raw.strip()
    # Exact alias (case-insensitive)
    for key, pathway in PATHWAY_ALIASES.items():
        if key.strip().lower() == r.lower():
            return pathway
    # Partial: if raw contains "business" -> Business Mgmt, "media" -> Media, "computing" -> Computing
    lower = r.lower()
    if "business" in lower or "biz" in lower or "mgmt" in lower or "management" in lower:
        return "Business Mgmt"
    if "media" in lower:
        return "Media"
    if "computing" in lower or "computer" in lower:
        return "Computing"
    return None

# Students to exclude from the system (by display name, case-insensitive)
EXCLUDED_STUDENT_NAMES = ["Cici", "Vivian"]

# Override display name by student_code (e.g. Excel had "Iris Yang", preferred name is "Yang yiqi")
NAME_OVERRIDES = {"51111798": "Yang yiqi", "51111738": "Chen Yu"}

# Excel column index (0-based) for each subject's current grade.
# Use component columns that actually contain data (Total columns 45, 61, 76... are often empty).
# AES columns are skipped on import so staff enter IELTS manually.
EXCEL_GRADE_COLUMNS = {
    "AES": 20,
    "AES Listening": 16,
    "AES Reading": 18,
    "AES Writing": 19,
    "AES Speaking": 17,
    "Business Management": 37,   # Individual Report (50%) in Business Management block
    "Mathematics": 52,           # Unseen 1 (40%) in Mathematics block
    "Statistics with Project Skills": 68,  # Unseen in Statistics block
    "Economics": 82,            # Unseen (50%) in Economics block
    "Physics": 97,              # Unseen (50%) in Physics block
    "Media Analysis": 110,      # Presentation (40%) / or 114 Advertising Campaign
    "Film Production": None,
    "Print & Digital News": None,
}


def _cell_has_grade(row: list, col: int) -> bool:
    """True if row has a numeric (or non-empty) grade at column col."""
    if col < 0 or col >= len(row):
        return False
    val = row[col]
    if val is None:
        return False
    if isinstance(val, str) and not val.strip():
        return False
    try:
        float(str(val).strip().rstrip("%"))
        return True
    except (TypeError, ValueError):
        return bool(str(val).strip())


def _format_grade_value(val) -> str:
    """Turn a cell value into a display grade (e.g. 65 -> '65%', 6.5 -> 'IELTS 6.5'). All numbers to 1 dp."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return ""
    if isinstance(val, str):
        val = val.strip().rstrip("%")
        try:
            val = float(val)
        except ValueError:
            return val.strip()
    try:
        v = float(val)
    except (TypeError, ValueError):
        return str(val).strip()
    v = round(v, 1)
    if v <= 9.5:
        return f"IELTS {v:.1f}"
    return f"{v:.1f}%"


def load_grades_excel(path: Optional[str] = None) -> list[dict]:
    """Load Excel and return list of student rows (all pathways)."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required; pip install openpyxl")

    excel_path = path or os.environ.get(
        "BELLERBYS_GRADES_EXCEL",
        os.path.join(os.path.dirname(__file__), "BNBU SAPM - Semester 1 Grades_v2.xlsx"),
    )
    if not os.path.isfile(excel_path):
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_name = "Sep 2025" if "Sep 2025" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Find the header row (Student ID in col 0)
    header_idx = next(
        (i for i, row in enumerate(rows[:15]) if row and row[0] == "Student ID"),
        0,
    )

    out = []
    for row in rows[header_idx + 1:]:
        if not row or row[STUDENT_ID_COL] is None:
            continue
        try:
            sid = int(row[STUDENT_ID_COL])
        except (TypeError, ValueError):
            continue

        pathway_raw = (row[PATHWAY_COL] or "").strip()
        pathway = _pathway_from_raw(pathway_raw)
        if pathway is None or pathway not in PATHWAY_SUBJECTS:
            continue

        first   = (row[FIRST_NAME_COL]   or "").strip()
        last    = (row[LAST_NAME_COL]    or "").strip()
        english = (row[ENGLISH_NAME_COL] or "").strip()
        student_name = english or f"{first} {last}".strip() or str(sid)
        if student_name.strip().lower() in [n.strip().lower() for n in EXCLUDED_STUDENT_NAMES if n]:
            continue

        subjects = list(PATHWAY_SUBJECTS[pathway])
        if pathway == "Business Mgmt":
            # Some Business students take Mathematics instead of Economics; infer from Excel row
            math_col = EXCEL_GRADE_COLUMNS["Mathematics"]
            econ_col = EXCEL_GRADE_COLUMNS["Economics"]
            has_math = _cell_has_grade(row, math_col)
            has_econ = _cell_has_grade(row, econ_col)
            if has_math and not has_econ:
                subjects = ["AES", "Business Management", "Statistics with Project Skills", "Mathematics"]
            else:
                subjects = ["AES", "Business Management", "Statistics with Project Skills", "Economics"]

        out.append({
            "student_code": str(sid),
            "student_name": student_name,
            "first_name":   first,
            "last_name":    last,
            "english_name": english,
            "pathway":      pathway,
            "subjects":     subjects,
            "grades":       {},
        })

    # Ensure unique display names so offer matching doesn't combine two people (e.g. Iris Yang vs Iris Chen)
    norm_counts: Counter = Counter((s["student_name"] or "").strip().lower() for s in out)
    for s in out:
        norm = (s["student_name"] or "").strip().lower()
        if norm and norm_counts.get(norm, 0) > 1:
            # Disambiguate: use First Last so "Iris" + Yang/Chen -> "Iris Yang" / "Iris Chen"
            full = f"{s['first_name']} {s['last_name']}".strip()
            if full:
                s["student_name"] = full
            else:
                s["student_name"] = f"{s['student_name']} ({s['student_code']})"
    # Apply name overrides (e.g. 51111798 -> "Yang yiqi")
    for s in out:
        if s["student_code"] in NAME_OVERRIDES:
            s["student_name"] = NAME_OVERRIDES[s["student_code"]]
    return out


def get_student_name_by_code(path: Optional[str] = None) -> dict[str, str]:
    """Return a dict mapping student_code (e.g. '12345') to the display student_name used in the app.
    Used to resolve student name from ID in PDF filename."""
    students = load_grades_excel(path)
    return {s["student_code"]: s["student_name"] for s in students}


def get_excluded_student_codes(path: Optional[str] = None) -> list[str]:
    """Return student_code for any student whose name is in EXCLUDED_STUDENT_NAMES (for DB cleanup)."""
    try:
        all_rows = load_grades_excel(path)
    except Exception:
        return []
    # load_grades_excel already filters them out, so we need to read Excel again without that filter
    excel_path = path or os.environ.get(
        "BELLERBYS_GRADES_EXCEL",
        os.path.join(os.path.dirname(__file__), "BNBU SAPM - Semester 1 Grades_v2.xlsx"),
    )
    if not os.path.isfile(excel_path):
        return []
    try:
        import openpyxl
    except ImportError:
        return []
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_name = "Sep 2025" if "Sep 2025" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header_idx = next(
        (i for i, row in enumerate(rows[:15]) if row and row[0] == "Student ID"),
        0,
    )
    excluded_lower = [n.strip().lower() for n in EXCLUDED_STUDENT_NAMES if n]
    codes = []
    for row in rows[header_idx + 1:]:
        if not row or row[STUDENT_ID_COL] is None:
            continue
        try:
            sid = str(int(row[STUDENT_ID_COL]))
        except (TypeError, ValueError):
            continue
        pathway_raw = (row[PATHWAY_COL] or "").strip()
        pathway = _pathway_from_raw(pathway_raw)
        if pathway is None or pathway not in PATHWAY_SUBJECTS:
            continue
        first = (row[FIRST_NAME_COL] or "").strip()
        last = (row[LAST_NAME_COL] or "").strip()
        english = (row[ENGLISH_NAME_COL] or "").strip()
        student_name = english or f"{first} {last}".strip() or sid
        if student_name.strip().lower() in excluded_lower:
            codes.append(sid)
    return codes


def load_grades_excel_with_grades(path: Optional[str] = None) -> list[dict]:
    """Load Excel and return list of student rows with grades dict filled from Excel columns."""
    students = load_grades_excel(path)
    if not students:
        return students

    excel_path = path or os.path.environ.get(
        "BELLERBYS_GRADES_EXCEL",
        os.path.join(os.path.dirname(__file__), "BNBU SAPM - Semester 1 Grades_v2.xlsx"),
    )
    if not os.path.isfile(excel_path):
        return students

    try:
        import openpyxl
    except ImportError:
        return students

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_name = "Sep 2025" if "Sep 2025" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = next(
        (i for i, row in enumerate(rows[:15]) if row and row[0] == "Student ID"),
        0,
    )

    # Build student_code -> row index (first match) so we can fill grades from same row
    code_to_row: dict[str, int] = {}
    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
        if not row or row[STUDENT_ID_COL] is None:
            continue
        try:
            sid = str(int(row[STUDENT_ID_COL]))
        except (TypeError, ValueError):
            continue
        pathway_raw = (row[PATHWAY_COL] or "").strip()
        pathway = _pathway_from_raw(pathway_raw)
        if pathway is None or pathway not in PATHWAY_SUBJECTS:
            continue
        if sid not in code_to_row:
            code_to_row[sid] = i

    for s in students:
        code = s["student_code"]
        row_idx = code_to_row.get(code)
        if row_idx is None:
            continue
        row = rows[row_idx]
        grades = {}
        for subject, col in EXCEL_GRADE_COLUMNS.items():
            if col is None:
                continue
            # Leave AES/IELTS empty on import; staff enter manually
            if subject == "AES" or subject.startswith("AES "):
                continue
            if col >= len(row):
                continue
            val = row[col]
            if val is None or (isinstance(val, str) and not val.strip()):
                continue
            try:
                _ = float(str(val).strip().rstrip("%"))
            except ValueError:
                continue
            grades[subject] = _format_grade_value(val)
        s["grades"] = grades

    return students


# Canonical pathway keys (used so grouping never drops a row due to case/whitespace)
_PATHWAY_CANONICAL: dict[str, str] = {
    "business mgmt": "Business Mgmt",
    "media": "Media",
    "computing": "Computing",
}


def get_grades_by_pathway(path: Optional[str] = None) -> dict[str, list[dict]]:
    """Return students grouped by pathway. Total count = sum of all list lengths."""
    all_rows = load_grades_excel(path)
    by_pathway: dict[str, list[dict]] = {"Business Mgmt": [], "Media": [], "Computing": []}
    for r in all_rows:
        raw = (r.get("pathway") or "").strip()
        key = _PATHWAY_CANONICAL.get(raw.lower()) if raw else None
        if key:
            by_pathway[key].append(r)
    return by_pathway
